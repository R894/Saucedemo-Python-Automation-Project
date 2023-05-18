import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    return sheet.cell(row=row_num, column=column_num).value


def get_rows_as_list(file, sheet_name):
    data_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        row = []
        for c in range(1, cols+1):
            row.append(sheet.cell(row=r, column=c).value)
        data_list.append(row)
    return data_list


def write_data(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column= column_num).value = data
    workbook.save(file)


